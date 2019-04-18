# Package openpyxl (help to operate .excel)
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

# Package numpy
import numpy as np

def data_init():
    file=load_workbook('data.xlsx')
    sheet_name_list = file.sheetnames
    sheet = file[sheet_name_list[0]]
    with open("scc.in", "r") as origin_file:
        data_list=origin_file.readlines()
        length=len(data_list)
        print(length)
        for i in range(1,length,1):
            sheet.cell(i + 1, 1).value = eval(data_list[i].split()[0])
            sheet.cell(i + 1, 2).value = eval(data_list[i].split()[1])

        origin_file.close()
    file.save('data.xlsx')

def main():
    data_init()

if __name__ == '__main__':
    main()
